"""Report endpoints: per-person summary and CSV export."""
import csv
import io
from datetime import date
from statistics import mean

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from ..database import get_db
from ..deps import get_app_tz, get_current_user, require_admin
from ..models import AttendanceRecord, Person, User
from ..schemas import PersonSummary, SummaryResponse

router = APIRouter(prefix="/api/reports", tags=["reports"])


@router.get("/summary", response_model=SummaryResponse)
async def summary(
    date_from: date,
    date_to: date,
    group: str | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> SummaryResponse:
    """Per-person attendance summary for a date range (inclusive).

    Admins see everyone (optionally filtered by group); regular users see
    their own summary.
    """
    if date_from > date_to:
        raise HTTPException(status_code=400, detail="date_from must be <= date_to")

    # People to summarize
    person_q = select(Person).where(Person.is_active.is_(True))
    if user.role != "admin":
        if user.person_id is None:
            return SummaryResponse(
                date_from=date_from, date_to=date_to, total_days=0, total_people=0, people=[]
            )
        person_q = person_q.where(Person.id == user.person_id)
    if group:
        person_q = person_q.where(Person.group_name == group)
    people = list((await db.execute(person_q.order_by(Person.full_name))).scalars().all())

    # Records in range for those people
    rec_q = (
        select(AttendanceRecord)
        .where(AttendanceRecord.date >= date_from, AttendanceRecord.date <= date_to)
        .order_by(AttendanceRecord.date)
    )
    if user.role != "admin" and user.person_id is not None:
        rec_q = rec_q.where(AttendanceRecord.person_id == user.person_id)
    records = list((await db.execute(rec_q)).scalars().all())

    total_days = (date_to - date_from).days + 1
    results: list[PersonSummary] = []
    by_person: dict[int, list[AttendanceRecord]] = {}
    for r in records:
        by_person.setdefault(r.person_id, []).append(r)

    for person in people:
        recs = by_person.get(person.id, [])
        present = sum(1 for r in recs if r.status == "present")
        absent = sum(1 for r in recs if r.status == "absent")
        half = sum(1 for r in recs if r.status == "half_day")
        holiday = sum(1 for r in recs if r.status == "holiday")
        unrecorded = max(total_days - len(recs), 0)

        durations = [
            int((r.check_out - r.check_in).total_seconds() // 60)
            for r in recs
            if r.check_in and r.check_out and r.check_out > r.check_in
        ]
        total_minutes = sum(durations) if durations else 0
        avg_minutes = int(mean(durations)) if durations else None

        in_times = [r.check_in for r in recs if r.check_in]
        out_times = [r.check_out for r in recs if r.check_out]
        app_tz = get_app_tz()

        def _avg_time(times) -> str | None:
            if not times:
                return None
            # average wall-clock time in the application timezone
            secs = mean(
                [
                    t.astimezone(app_tz).hour * 3600
                    + t.astimezone(app_tz).minute * 60
                    + t.astimezone(app_tz).second
                    for t in times
                ]
            )
            hh, mm = int(secs // 3600), int((secs % 3600) // 60)
            return f"{hh:02d}:{mm:02d}"

        avg_in = _avg_time(in_times)
        avg_out = _avg_time(out_times)

        rate = round((present + half) / total_days, 3) if total_days else 0.0
        results.append(
            PersonSummary(
                person_id=person.id,
                person_name=person.full_name,
                group_name=person.group_name,
                total_days=total_days,
                present_days=present,
                absent_days=absent,
                half_days=half,
                holiday_days=holiday,
                unrecorded_days=unrecorded,
                total_work_minutes=total_minutes,
                avg_work_minutes=avg_minutes,
                avg_check_in=avg_in,
                avg_check_out=avg_out,
                attendance_rate=rate,
            )
        )

    return SummaryResponse(
        date_from=date_from,
        date_to=date_to,
        total_days=total_days,
        total_people=len(results),
        people=results,
    )


@router.get("/export")
async def export_csv(
    date_from: date,
    date_to: date,
    group: str | None = None,
    person_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
):
    """Download attendance records as CSV for a date range."""
    query = (
        select(AttendanceRecord)
        .options(joinedload(AttendanceRecord.person))
        .join(Person)
        .where(AttendanceRecord.date >= date_from, AttendanceRecord.date <= date_to)
        .order_by(AttendanceRecord.date, Person.full_name)
    )
    if group:
        query = query.where(Person.group_name == group)
    if person_id:
        query = query.where(AttendanceRecord.person_id == person_id)
    records = list((await db.execute(query)).scalars().all())

    buf = io.StringIO()
    writer = csv.writer(buf)
    app_tz = get_app_tz()
    writer.writerow(
        ["date", "person", "group", "status", "check_in", "check_out", "duration_minutes", "note"]
    )
    for r in records:
        dur = ""
        if r.check_in and r.check_out:
            dur = int((r.check_out - r.check_in).total_seconds() // 60)
        writer.writerow(
            [
                r.date.isoformat(),
                r.person.full_name if r.person else "",
                r.person.group_name if r.person else "",
                r.status,
                r.check_in.astimezone(app_tz).strftime("%Y-%m-%d %H:%M") if r.check_in else "",
                r.check_out.astimezone(app_tz).strftime("%Y-%m-%d %H:%M") if r.check_out else "",
                dur,
                r.note or "",
            ]
        )
    buf.seek(0)
    filename = f"attendance_{date_from}_{date_to}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/xlsx")
async def export_xlsx(
    date_from: date,
    date_to: date,
    group: str | None = None,
    person_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
):
    """Download attendance records as Excel for a date range."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = (
        select(AttendanceRecord)
        .options(joinedload(AttendanceRecord.person))
        .join(Person)
        .where(AttendanceRecord.date >= date_from, AttendanceRecord.date <= date_to)
        .order_by(AttendanceRecord.date, Person.full_name)
    )
    if group:
        query = query.where(Person.group_name == group)
    if person_id:
        query = query.where(AttendanceRecord.person_id == person_id)
    records = list((await db.execute(query)).scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Date", "Person", "Group", "Status", "Check In", "Check Out", "Duration (min)", "Note"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    status_colors = {
        "present": "22c55e", "absent": "ef4444",
        "half_day": "f59e0b", "holiday": "3b82f6",
    }
    app_tz = get_app_tz()

    for row_idx, r in enumerate(records, 2):
        dur = ""
        if r.check_in and r.check_out:
            dur = int((r.check_out - r.check_in).total_seconds() // 60)
        values = [
            r.date.isoformat(),
            r.person.full_name if r.person else "",
            r.person.group_name if r.person else "",
            r.status,
            r.check_in.astimezone(app_tz).strftime("%Y-%m-%d %H:%M") if r.check_in else "",
            r.check_out.astimezone(app_tz).strftime("%Y-%m-%d %H:%M") if r.check_out else "",
            dur,
            r.note or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            if col == 4 and val in status_colors:
                cell.font = Font(color=status_colors[val], bold=True)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"attendance_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
